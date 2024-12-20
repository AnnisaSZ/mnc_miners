from odoo import models, fields, api, exceptions, _
from odoo.exceptions import ValidationError
from odoo.tools import format_datetime
from datetime import datetime, timedelta
from collections import defaultdict

import base64
from io import BytesIO
import qrcode
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from ..controllers.main import float_to_time_str

# import itertools
import requests
import json

LIST_CI = [
    ('CI_Ontime', 'Check In Ontime'),
    ('Late', 'Late'),
    ('Alpha', 'Alpha'),
    ('Permission', 'Permission'),
    ('Sick', 'Sick'),
    ('Leave', 'Leave'),
]

LIST_CO = [
    ('Alpha', 'Alpha'),
    ('CO_Ontime', 'Check Out Ontime'),
    ('Overtime', 'Overtime'),
    ('Early_CO', 'Early Check Out'),
    ('Leave', 'On Leave'),
    ('Sick', 'Sick'),
]


class HRAttendance(models.Model):
    _inherit = 'hr.attendance'

    employee_id = fields.Many2one('hr.employee', string="Not Used", default=False, required=False, ondelete='cascade', index=True)
    company_id = fields.Many2one('res.company', string="Company", related='mncei_employee_id.company', store=True)
    check_in = fields.Datetime(string="Check In", default=fields.Datetime.now(), required=False, store=True)
    mncei_employee_id = fields.Many2one('mncei.employee', string="Employee", default=False, required=True, ondelete='cascade', index=True, store=True)
    resouce_id = fields.Many2one('resource.calendar', compute='_get_type_attendance', string="Working Time", store=True)
    resouce_line_id = fields.Many2one('resource.calendar.attendance', compute='_get_type_attendance', string="Working Time Line", store=True)
    department_id = fields.Many2one(
        'mncei.department',
        string='Department', related='mncei_employee_id.department'
    )

    img_check_in = fields.Binary(string='Image Check In', store=True, attachment=True)
    img_check_out = fields.Binary(string='Image Check Out', store=True, attachment=True)

    remarks = fields.Text("Remarks", store=True)
    overtime = fields.Float('Overtime', compute='_get_overtime', store=True)

    # Check In
    check_in_id = fields.Integer('ID Check IN', store=True)
    type_ci = fields.Selection(LIST_CI, compute='get_type_ci', string="Notes CI", store=True)
    location_ci_id = fields.Many2one(
        'res.att.location',
        string='Location', store=True
    )
    location_ci_notes = fields.Text("Notes Location", store=True)
    location_ci_point = fields.Text("Point Location", store=True)
    location_ci_other = fields.Boolean("Other Location", related="location_ci_id.is_other")
    remarks_ci = fields.Text("Remarks", store=True)

    # Check Out
    check_out_id = fields.Integer('ID Check Out', store=True)
    type_co = fields.Selection(LIST_CO, compute='get_type_co', string="Notes CO", store=True)
    location_co_id = fields.Many2one(
        'res.att.location',
        string='Location', store=True
    )
    location_co_notes = fields.Text("Notes Location", store=True)
    location_co_point = fields.Text("Point Location", store=True)
    location_co_other = fields.Boolean("Other Location", related="location_co_id.is_other")
    remarks_co = fields.Text("Remarks", store=True)

    attendance_type = fields.Selection([
        ('Normal', 'Normal'),
        ('Late', 'Late'),
        ('Early_CO', 'Early Check Out'),
        ('Alpha', 'Alpha'),
        ('Overtime', 'Overtime'),
        ('Permission', 'Permission'),
        ('Sick', 'Sick'),
    ], default='Normal', string="Notes", store=True)

    # CR 2 HR
    leave_type_id = fields.Many2one(
        'hr.leave.type', string='Type Details'
    )
    alpha_details = fields.Selection([
        ('Leave', 'Leave'),
        ('Sick', 'Sick'),
    ], string="Details", store=True)
    # type_details_id = fields.Many2one('hr.leave.type.details', string='Details')
    document = fields.Binary(
        string='Attachment',
        attachment=True,
    )
    document_name = fields.Char('Document Name', store=True)
    reason = fields.Text("Reason")
    is_leave = fields.Boolean('Is Leave', store=True)
    is_sick = fields.Boolean('Is Sick', store=True)
    is_alpha = fields.Boolean('Is Alpha', store=True)

    def export_to_excel(self, records):
        attendance_obj = self.env['hr.attendance']
        # Buat workbook dan worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Datas Attendance'

        # Tuliskan beberapa data contoh
        center_alignment = Alignment(horizontal='center', vertical='center')
        headers = ['ID', 'Employee Name', 'Department', 'Check In', 'Tipe In', 'Check Out', 'Tipe Out', 'Worked Hours', 'Overtime']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.alignment = center_alignment
            column_width = len(header) + 2  # Menambahkan beberapa spasi ekstra untuk padding
            worksheet.column_dimensions[get_column_letter(col_num)].width = column_width
        row = 2
        for res_id in records:
            record = self.browse(res_id)
            # record = res_id
            worksheet.cell(row=row, column=1, value=record.id)
            worksheet.cell(row=row, column=2, value=record.mncei_employee_id.nama_lengkap)
            worksheet.cell(row=row, column=3, value=record.department_id.name)
            check_in = worksheet.cell(row=row, column=4, value=(record.check_in + timedelta(hours=7)) if record.check_in else 'N/A')
            check_in.alignment = center_alignment
            worksheet.cell(row=row, column=5, value=dict(attendance_obj._fields['type_ci'].selection).get(record.type_ci))
            check_out = worksheet.cell(row=row, column=6, value=(record.check_out + timedelta(hours=7)) if record.check_out else 'N/A')
            check_out.alignment = center_alignment
            worksheet.cell(row=row, column=7, value=dict(attendance_obj._fields['type_co'].selection).get(record.type_co))
            worksheet.cell(row=row, column=8, value=float_to_time_str(record.worked_hours))
            worksheet.cell(row=row, column=9, value=float_to_time_str(record.overtime))
            # Department
            column_width = len(record.department_id.name) + 2 if (len("Department") + 2) < (len(record.department_id.name) + 2) else len("Department") + 2
            worksheet.column_dimensions[get_column_letter(3)].width = column_width
            # Check In
            if record.check_in:
                column_width = len(record.check_in.strftime('%Y-%m-%d %H:%M:%S')) + 2
                worksheet.column_dimensions[get_column_letter(4)].width = column_width
                worksheet.column_dimensions[get_column_letter(5)].width = column_width
            # Check Out
            if record.check_out:
                column_width = len(record.check_out.strftime('%Y-%m-%d %H:%M:%S')) + 2
                worksheet.column_dimensions[get_column_letter(6)].width = column_width
                worksheet.column_dimensions[get_column_letter(7)].width = column_width
            row += 1

        # Buat QR code dengan tanggal export
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=5,
            border=4,
        )
        export_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        user_id = self.env['res.users'].browse(self.env.uid)
        combination_values = ("%s %s") % (user_id.name, export_date)
        qr.add_data(combination_values)
        qr.make(fit=True)

        img = qr.make_image(fill='black', back_color='white')

        # Simpan QR code ke dalam buffer
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)

        # Buat gambar QR code untuk disisipkan dalam Excel
        qr_image = OpenpyxlImage(buffer)
        worksheet.add_image(qr_image, 'A{}'.format(row + 2))

        # Simpan workbook ke buffer
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        # Buat attachment Odoo
        attachment = self.env['ir.attachment'].create({
            'name': 'Export_with_QRCode.xlsx',
            'type': 'binary',
            'public': True,
            'datas': base64.b64encode(excel_buffer.read()),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })
        return attachment.id

    def get_attachments_link(self, records):
        attachment_id = self.export_to_excel(records)
        return '/web/content/%s?download=true' % (str(attachment_id))

    @api.onchange('alpha_details')
    def change_reason_alpha(self):
        if self.alpha_details:
            if self.alpha_details == 'Sick':
                self.is_sick = True
                self.is_leave = False
            elif self.alpha_details == 'Leave':
                self.is_leave = True
                self.is_sick = False
        else:
            self.is_sick = False
            self.is_leave = False

    @api.depends('check_in', 'resouce_id', 'mncei_employee_id', 'resouce_line_id', 'is_alpha', 'is_leave', 'is_sick')
    def get_type_ci(self):
        for attendance in self:
            if attendance.resouce_line_id:
                if attendance.is_alpha:
                    if attendance.is_leave:
                        attendance.type_ci = 'Leave'
                    elif attendance.is_sick:
                        attendance.type_ci = 'Sick'
                    elif not attendance.is_leave and not attendance.is_sick:
                        attendance.type_ci = 'Alpha'
                elif attendance.check_in and attendance.mncei_employee_id:
                    result = attendance.set_type_attendance(attendance.resouce_id, "check_in", attendance.resouce_line_id)
                    attendance.type_ci = result
            elif not attendance.check_in:
                attendance.type_ci = ''
            else:
                attendance.type_ci = 'CI_Ontime'

    @api.depends('check_out', 'resouce_id', 'mncei_employee_id', 'resouce_line_id', 'is_alpha', 'is_leave', 'is_sick')
    def get_type_co(self):
        for attendance in self:
            if attendance.resouce_line_id:
                if attendance.is_alpha:
                    if attendance.is_leave:
                        attendance.type_co = 'Leave'
                    elif attendance.is_sick:
                        attendance.type_co = 'Sick'
                    else:
                        attendance.type_co = 'Alpha'
                elif attendance.check_out and attendance.mncei_employee_id:
                    result = attendance.set_type_attendance(attendance.resouce_id, "check_out", attendance.resouce_line_id)
                    attendance.type_co = result
            elif not attendance.check_out:
                attendance.type_co = ''
            else:
                attendance.type_co = ''

    @api.depends('check_in', 'check_out', 'mncei_employee_id', 'mncei_employee_id.working_time_id', 'mncei_employee_id.shift_temp_ids')
    def _get_type_attendance(self):
        for attendance in self:
            working_time = attendance.mncei_employee_id.working_time_id
            time_get = fields.Date.today()
            working_line = False
            if attendance.check_in:
                check_in = attendance.check_in + timedelta(hours=7)
                time_get = check_in.date()
                dayofweek = check_in.strftime("%A")
                resouce_line_id = working_time.attendance_ids.filtered(lambda x: dict(self.env['resource.calendar.attendance']._fields['dayofweek'].selection).get(x.dayofweek) == dayofweek)
                working_line = resouce_line_id
            else:
                check_out = attendance.check_out + timedelta(hours=7)
                time_get = check_out.date()
                dayofweek = check_out.strftime("%A")
                resouce_line_id = working_time.attendance_ids.filtered(lambda x: dict(self.env['resource.calendar.attendance']._fields['dayofweek'].selection).get(x.dayofweek) == dayofweek)
                working_line = resouce_line_id

            if attendance.mncei_employee_id.shift_temp_ids:
                shift_id = attendance.mncei_employee_id.shift_temp_ids.filtered(lambda x: x.start_date <= time_get and x.end_date >= time_get)
                if shift_id:
                    working_time = shift_id.working_time_id
                    if working_time:
                        dayofweek = attendance.check_in.strftime("%A")
                        resouce_line_id = working_time.attendance_ids.filtered(lambda x: dict(self.env['resource.calendar.attendance']._fields['dayofweek'].selection).get(x.dayofweek) == dayofweek)
                        working_line = resouce_line_id
            #     else:
            #         if attendance.check_in:
            #             check_in = attendance.check_in + timedelta(hours=7)
            #             dayofweek = check_in.strftime("%A")
            #             resouce_line_id = working_time.attendance_ids.filtered(lambda x: dict(self.env['resource.calendar.attendance']._fields['dayofweek'].selection).get(x.dayofweek) == dayofweek)
            #             working_line = resouce_line_id
            #         else:
            #             check_out = attendance.check_out + timedelta(hours=7)
            #             dayofweek = check_out.strftime("%A")
            #             resouce_line_id = working_time.attendance_ids.filtered(lambda x: dict(self.env['resource.calendar.attendance']._fields['dayofweek'].selection).get(x.dayofweek) == dayofweek)
            #             working_line = resouce_line_id
            # else:
            #     if attendance.check_in:
            #         check_in = attendance.check_in + timedelta(hours=7)
            #         dayofweek = check_in.strftime("%A")
            #         resouce_line_id = working_time.attendance_ids.filtered(lambda x: dict(self.env['resource.calendar.attendance']._fields['dayofweek'].selection).get(x.dayofweek) == dayofweek)
            #         working_line = resouce_line_id
            #     else:
            #         check_out = attendance.check_out + timedelta(hours=7)
            #         dayofweek = check_out.strftime("%A")
            #         resouce_line_id = working_time.attendance_ids.filtered(lambda x: dict(self.env['resource.calendar.attendance']._fields['dayofweek'].selection).get(x.dayofweek) == dayofweek)
            #         working_line = resouce_line_id
            # Set Values
            attendance.resouce_id = working_time
            attendance.resouce_line_id = working_line

    def set_type_attendance(self, working_time, ttype, resouce_line_id=False):
        if ttype == "check_in":
            check_in = self.check_in + timedelta(hours=7)
            ci_time = check_in.time()
            if resouce_line_id:
                result = self._check_late(working_time, ci_time, type=ttype, resouce_line=resouce_line_id)
            else:
                result = self._check_late(working_time, ci_time, type=ttype)
            if result:
                return result
        if ttype == "check_out":
            check_out = self.check_out + timedelta(hours=7)
            co_time = check_out.time()
            if resouce_line_id:
                result = self._check_late(working_time, co_time, type=ttype, resouce_line=resouce_line_id)
            else:
                result = self._check_late(working_time, co_time, type=ttype)
            if result:
                return result

    def _check_late(self, working_time, time, type=False, resouce_line=False):
        # Parameter
        # variable time_float >> time actual
        if resouce_line:
            work_ci = resouce_line.hour_from
            work_co = resouce_line.hour_to
            work_ot = resouce_line.start_overtime
        else:
            if working_time.attendance_ids:
                work_ci = working_time.attendance_ids[0].hour_from
                work_co = working_time.attendance_ids[0].hour_to
                work_ot = working_time.attendance_ids[0].start_overtime
        # Check Atteandance
        time_float = time.hour + time.minute / 60.0
        if type == 'check_in':
            if (work_ci + working_time.limit_attendance) < time_float:
                return 'Late'
            elif (work_ci + working_time.limit_attendance) >= time_float:
                return 'CI_Ontime'
        if type == 'check_out':
            if work_co > time_float:
                return 'Early_CO'
            elif work_ot < time_float and work_ot != 0.0:
                return 'Overtime'
            elif work_co <= time_float:
                return 'CO_Ontime'

    @api.depends('check_out', 'worked_hours', 'resouce_id')
    def _get_overtime(self):
        for attendance in self:
            # print("XXXXXXXXXXXXXXXXXXXXXX")
            # print(attendance.worked_hours)
            # Fungsi dijalankan jika menggunakan kondisi harus menghitung sebelum dan sesudah jam masuk
            # if attendance.resouce_line_id:
            #     work_ci = attendance.resouce_line_id.hour_from
            #     work_co = attendance.resouce_line_id.hour_to
            # else:
            #     if attendance.resouce_id.attendance_ids:
            #         work_ci = attendance.resouce_id.attendance_ids[0].hour_from
            #         work_co = attendance.resouce_id.attendance_ids[0].hour_to
            # print(work_ci)
            # print(work_co)
            calc_ot = attendance.worked_hours - attendance.resouce_id.hours_per_day
            attendance.overtime = calc_ot if calc_ot > 0.0 else 0.0

    # By Working Time
    @api.model
    def action_get_transaction_by_location(self, location_id):
        working_time_ids = self.env['resource.calendar'].search([('loc_working_id', '=', location_id)])
        if working_time_ids:

            start = '%s 00:00:00' % (fields.Date.today())
            end = '%s 23:00:00' % (fields.Date.today())
            # Loop
            for working_time in working_time_ids:
                if working_time.is_shift:
                    for shift_id in working_time.shift_temp_ids.filtered(lambda x: x.start_date <= fields.Date.today() and x.end_date >= fields.Date.today()):
                        employee_code = shift_id.mncei_employee_id.wdms_code
                        if employee_code:
                            start = '%s 00:00:00' % (fields.Date.today() - timedelta(days=1))
                            datas = {
                                'start_date': start,
                                'end_date': end,
                                'employee_code': employee_code,
                                'is_shift': True
                            }
                            delayed = (
                                self.env["queue.job"]
                                .with_delay(
                                    priority=None,
                                    max_retries=None,
                                    channel=None,
                                    description="Attendance",
                                )
                                ._create_attendance(datas)
                            )
                        # self.action_get_transaction(start, end, employee_code, is_shift=True)
                else:
                    for employee_id in working_time.mncei_employee_id:
                        employee_code = employee_id.wdms_code
                        if employee_code:
                            datas = {
                                'start_date': start,
                                'end_date': end,
                                'employee_code': employee_code,
                                'is_shift': False
                            }
                            delayed = (
                                self.env["queue.job"]
                                .with_delay(
                                    priority=None,
                                    max_retries=None,
                                    channel=None,
                                    description="Attendance",
                                )
                                ._create_attendance(datas)
                            )
                        # self.action_get_transaction(start, end, employee_code)

    @api.model
    def action_get_transaction(self, start_date, end_date, employee_code, is_shift=False):
        config_id = self.env['wdms.config'].search([('token', '!=', False)], limit=1)
        if config_id:
            # start = '%s 00:00:00' % (fields.Date.today())
            # end = '%s 23:00:00' % (fields.Date.today())
            start = start_date
            end = end_date
            employee_code = employee_code
            url = _("%s/iclock/api/transactions/?start_time=%s&end_time=%s&emp_code=%s") % (config_id.ip_link, start, end, employee_code)
            auth = _("JWT %s") % (config_id.token)

            payload = {}
            headers = {
                'Content-Type': 'application/json',
                'Authorization': auth
            }

            response = requests.request("GET", url, headers=headers, data=payload)
            # Check Result
            data = json.loads(response.text)
            if response.status_code == 200:
                next = False
                grouped_data = {}
                if len(data['data']) > 0:
                    if data['next']:
                        link_next = data['next']
                        next = True
                        # Get Grouping Data
                        grouped_data = self.group_data_shift(data['data'], grouped_data)
                    else:
                        grouped_data = self.group_data_shift(data['data'], grouped_data)
                    while next:
                        if link_next:
                            next, link_next, data = self.check_next_page(link_next, headers, grouped_data)
                            grouped_data = data
                    self.create_attendance(grouped_data)
                    return
            else:
                raise ValidationError(_("Get Token Failed"))
            return

        return

    def check_next_page(self, url, headers, grouped_data):
        next = False
        next_page = False
        response = requests.request("GET", url, headers=headers, data={})
        data = json.loads(response.text)
        if response.status_code == 200:
            if data['next']:
                next_page = data['next']
                grouped_data = self.group_data_shift(data['data'], grouped_data)
                # grouped_data = self.group_datas(data['data'], grouped_data)
                next = True
            else:
                next_page = False
                next = False
        else:
            raise ValidationError(_("Get Token Failed"))
        return next, next_page, grouped_data

    def group_datas(self, datas, grouped_data):
        grouped_data = grouped_data

        for item in datas:
            emp_code = item["emp_code"]
            punch_state_display = item["punch_state_display"]
            punch_time = item["punch_time"]

            if emp_code not in grouped_data:
                grouped_data[emp_code] = {punch_state_display: {"punch_time": punch_time, "data": item}}
            elif punch_state_display not in grouped_data[emp_code]:
                grouped_data[emp_code][punch_state_display] = {"punch_time": punch_time, "data": item}
            else:
                current_punch_time = grouped_data[emp_code][punch_state_display]["punch_time"]
                if punch_time < current_punch_time:
                    grouped_data[emp_code][punch_state_display] = {"punch_time": punch_time, "data": item}

        return grouped_data

    def group_data_shift(self, datas, grouped_data):
        grouped_data = {}
        # Mengelompokkan data berdasarkan emp_code
        for entry in datas:
            emp_code = entry["emp_code"]
            if emp_code not in grouped_data:
                grouped_data[emp_code] = {"check_ins": False, "check_outs": False}
            if entry["punch_state_display"] == "Check In":
                # grouped_data[emp_code]["check_ins"].append(datetime.strptime(entry["punch_time"], "%Y-%m-%d %H:%M:%S"))
                grouped_data[emp_code]["check_ins"] = datetime.strptime(entry["punch_time"], "%Y-%m-%d %H:%M:%S")
            elif entry["punch_state_display"] == "Check Out":
                # grouped_data[emp_code]["check_outs"].append(datetime.strptime(entry["punch_time"], "%Y-%m-%d %H:%M:%S"))
                grouped_data[emp_code]["check_outs"] = datetime.strptime(entry["punch_time"], "%Y-%m-%d %H:%M:%S")

        print(grouped_data)
        result = []
        for emp_code, punches in grouped_data.items():
            if punches["check_ins"] and punches["check_outs"]:
                check_in = punches["check_ins"]
                check_out = punches["check_outs"]
                result.append({
                    "emp_code": emp_code,
                    "check_in": check_in,
                    "check_out": check_out
                })
            elif punches["check_ins"] and not punches["check_outs"]:
                check_in = punches["check_ins"]
                result.append({
                    "emp_code": emp_code,
                    "check_in": check_in,
                    "check_out": False
                })
            elif punches["check_outs"] and not punches["check_ins"]:
                check_out = punches["check_outs"]
                result.append({
                    "emp_code": emp_code,
                    "check_in": False,
                    "check_out": check_out,
                })
        return result

    def create_attendance(self, datas):
        for data in datas:
            employee_id = self.env['mncei.employee'].search([('wdms_code', '=', data['emp_code'])], limit=1)
            location_id = self.env['res.att.location'].search([('is_other', '=', True)], limit=1) or False
            if employee_id:
                # Check In
                if data["check_in"]:
                    check_in = data["check_in"]
                    start, end = self.get_start_end_datetime(check_in.date())
                    start_time = self.str_to_datetime(start) - timedelta(hours=7)
                    end_time = self.str_to_datetime(end) - timedelta(hours=7)
                    attendance_id = self.env['hr.attendance'].search([('mncei_employee_id', '=', employee_id.id), ('check_in', '<=', end_time), ('check_in', '>=', start_time)])
                    if not attendance_id:
                        attendance_co_id = self.env['hr.attendance'].search([('mncei_employee_id', '=', employee_id.id), ('check_out', '<=', end_time), ('check_out', '>=', start_time)])
                        if attendance_co_id:
                            attendance_co_id.write({'check_in': check_in - timedelta(hours=7)})
                        else:
                            self.env['hr.attendance'].create({
                                'mncei_employee_id': employee_id.id,
                                'location_ci_id': location_id.id if location_id else False,
                                'remarks_ci': 'Mesin Absen',
                                'check_in': check_in - timedelta(hours=7)
                            })
                # Check Out
                if data["check_out"]:
                    check_out = data["check_out"]
                    if data['check_in']:
                        start_co, end_co = self.get_start_end_datetime(data['check_in'].date())
                    else:
                        start_co, end_co = self.get_start_end_datetime(check_out.date())
                    start_co_time = self.str_to_datetime(start_co) - timedelta(hours=7)
                    end_co_time = self.str_to_datetime(end_co) - timedelta(hours=7)
                    attendance_ci_id = self.env['hr.attendance'].search([('mncei_employee_id', '=', employee_id.id), ('check_in', '<=', end_co_time), ('check_in', '>=', start_co_time)])
                    if attendance_ci_id:
                        if not attendance_ci_id.check_out:
                            attendance_ci_id.write({
                                'check_out': check_out - timedelta(hours=7),
                                'location_co_id': location_id.id if location_id else False,
                                'remarks_co': 'Mesin Absen',
                            })
                    else:
                        attendance_co_id = self.env['hr.attendance'].search([('mncei_employee_id', '=', employee_id.id), ('check_out', '<=', end_co_time), ('check_out', '>=', start_co_time)])
                        if not attendance_co_id:
                            self.env['hr.attendance'].create({
                                'mncei_employee_id': employee_id.id,
                                'check_in': False,
                                'check_out': check_out - timedelta(hours=7),
                                'location_co_id': location_id.id if location_id else False,
                                'remarks_co': 'Mesin Absen',
                            })
        return

    def get_start_end_datetime(self, date):
        start = _('%s 00:00:00' % (date))
        end = _('%s 23:59:00' % (date))
        return start, end

    def str_to_datetime(self, time_char):
        return datetime.strptime(time_char, "%Y-%m-%d %H:%M:%S")

    @api.constrains('check_in', 'check_out')
    def _check_validity_check_in_check_out(self):
        """ verifies if check_in is earlier than check_out. """
        for attendance in self:
            if not attendance.mncei_employee_id:
                if attendance.check_in and attendance.check_out:
                    if attendance.check_out < attendance.check_in:
                        raise exceptions.ValidationError(_('"Check Out" time cannot be earlier than "Check In" time.'))

    # Override
    @api.constrains('check_in', 'check_out', 'employee_id', 'mncei_employee_id')
    def _check_validity(self):
        """ Verifies the validity of the attendance record compared to the others from the same employee.
            For the same employee we must have :
                * maximum 1 "open" attendance record (without check_out)
                * no overlapping time slices with previous employee records
        """
        for attendance in self:
            if not attendance.mncei_employee_id:
                # we take the latest attendance before our check_in time and check it doesn't overlap with ours
                last_attendance_before_check_in = self.env['hr.attendance'].search([
                    ('employee_id', '=', attendance.employee_id.id),
                    ('check_in', '<=', attendance.check_in),
                    ('id', '!=', attendance.id),
                ], order='check_in desc', limit=1)
                if last_attendance_before_check_in and last_attendance_before_check_in.check_out and last_attendance_before_check_in.check_out > attendance.check_in:
                    raise exceptions.ValidationError(_("Cannot create new attendance record for %(empl_name)s, the employee was already checked in on %(datetime)s") % {
                        'empl_name': attendance.employee_id.name,
                        'datetime': format_datetime(self.env, attendance.check_in, dt_format=False),
                    })

                if not attendance.check_out:
                    # if our attendance is "open" (no check_out), we verify there is no other "open" attendance
                    no_check_out_attendances = self.env['hr.attendance'].search([
                        ('employee_id', '=', attendance.employee_id.id),
                        ('check_out', '=', False),
                        ('id', '!=', attendance.id),
                    ], order='check_in desc', limit=1)
                    if no_check_out_attendances:
                        raise exceptions.ValidationError(_("Cannot create new attendance record for %(empl_name)s, the employee hasn't checked out since %(datetime)s") % {
                            'empl_name': attendance.employee_id.name,
                            'datetime': format_datetime(self.env, no_check_out_attendances.check_in, dt_format=False),
                        })
                else:
                    # we verify that the latest attendance with check_in time before our check_out time
                    # is the same as the one before our check_in time computed before, otherwise it overlaps
                    last_attendance_before_check_out = self.env['hr.attendance'].search([
                        ('employee_id', '=', attendance.employee_id.id),
                        ('check_in', '<', attendance.check_out),
                        ('id', '!=', attendance.id),
                    ], order='check_in desc', limit=1)
                    if last_attendance_before_check_out and last_attendance_before_check_in != last_attendance_before_check_out:
                        raise exceptions.ValidationError(_("Cannot create new attendance record for %(empl_name)s, the employee was already checked in on %(datetime)s") % {
                            'empl_name': attendance.employee_id.name,
                            'datetime': format_datetime(self.env, last_attendance_before_check_out.check_in, dt_format=False),
                        })

    # API
    def get_data_present(self, date, employee_ids):
        attendance_obj = self.env['hr.attendance']
        total = 0
        late = 0
        early_co = 0
        ot = 0
        for employee_id in employee_ids:
            start, end = attendance_obj.get_start_end_datetime(date)
            start_time = attendance_obj.str_to_datetime(start) - timedelta(hours=7)
            end_time = attendance_obj.str_to_datetime(end) - timedelta(hours=7)
            attendance_id = attendance_obj.search([('mncei_employee_id', '=', employee_id), ('check_in', '<=', end_time), ('check_in', '>=', start_time)])
            if attendance_id:
                total += 1
                if attendance_id.type_ci == "Late":
                    late += 1
                if attendance_id.type_co == "Early_CO":
                    early_co += 1
                if attendance_id.type_co == "Overtime":
                    ot += 1
        res = {
            'total': total,
            'late': late,
            'early_co': early_co,
            'ot': ot,
        }
        return res

    def get_attendance_id(self, employee_id, start_date, end_date):
        attendance_ci_id = self.env['hr.attendance'].search([('mncei_employee_id', '=', employee_id.id), ('check_in', '>=', start_date), ('check_in', '<=', end_date)], limit=1)
        attendance_co_id = self.env['hr.attendance'].search([('mncei_employee_id', '=', employee_id.id), ('check_out', '>=', start_date), ('check_out', '<=', end_date)], limit=1)

        if not attendance_co_id and not attendance_ci_id:
            return False
        else:
            if attendance_co_id:
                return attendance_co_id
            elif attendance_ci_id:
                return attendance_ci_id

    def to_create(self, date_time, employee_id, location_id, params):
        res_id = False
        res_data = {}
        date_format = "%Y-%m-%d"
        # current = params['date']
        # start, end = self.get_start_end_datetime(date_time.date())
        current = datetime.strptime(params['date'], date_format).date()
        start = datetime.combine(current, datetime.min.time()) - timedelta(hours=7)
        end = datetime.combine(current, datetime.max.time()) - timedelta(hours=7)
        ttype = params['type']
        # Parameter CI/CO
        if ttype == 'check_in':
            res_data.update({
                'location_ci_id': location_id.id,
                'location_ci_notes': params['location_notes'],
                'location_ci_point': params['location_point'],
                'remarks_ci': params['notes'],
                'check_in': date_time,
                'img_check_in': params['photo']
            })
        if ttype == 'check_out':
            res_data.update({
                'location_co_id': location_id.id,
                'location_co_notes': params['location_notes'],
                'location_co_point': params['location_point'],
                'remarks_co': params['notes'],
                'check_out': date_time,
                'img_check_out': params['photo']
            })
        # Get Attendance
        if 'id' in params:
            if params['id'] != 0:
                # if params['id'] != 0:
                attendance_id = self.browse(params['id'])
                if attendance_id:
                    res_id = attendance_id
                    res_id.write(res_data)
                else:
                    raise ValidationError(_("ID Not Found"))
            else:
                attendance_id = self.get_attendance_id(employee_id, start, end)
                if not attendance_id:
                    res_data.update({
                        'mncei_employee_id': employee_id.id
                    })
                    if ttype == 'check_out':
                        res_data['check_in'] = False
                    res_id = self.create(res_data)
                else:
                    res_id = attendance_id
                    res_id.write(res_data)
        else:
            attendance_id = self.get_attendance_id(employee_id, start, end)
            if not attendance_id:
                res_data.update({
                    'mncei_employee_id': employee_id.id
                })
                if ttype == 'check_out':
                    res_data['check_in'] = False
                res_id = self.create(res_data)
            else:
                res_id = attendance_id
                res_id.write(res_data)
        self.add_attachment(res_id, ttype, params['photo'])
        return res_id

    # -------- To Create Attachment ------
    def add_attachment(self, attendance_id, ttype, filename=False):
        attachment = self.env['ir.attachment'].create(
            {
                'name': str(attendance_id.id),
                'company_id': attendance_id.mncei_employee_id.company.id,
                'public': True,
                'type': 'binary',
                'datas': filename,
                'res_model': 'hr.attendance',
                'res_id': attendance_id.id
            })
        if ttype == 'check_in':
            attachment.write({
                'res_field': 'check_in'
            })
        elif ttype == 'check_out':
            attachment.write({
                'res_field': 'check_out'
            })
        return attachment

    # -------- Auto Create & Send Email Alpha ------
    @api.model
    def check_employee_attendance(self):
        today = fields.Date.today()
        attendance_obj = self.env['hr.attendance']
        leave_obj = self.env['hr.leave']
        employee_ids = self.env['mncei.employee'].search([('state', '=', 'verified'), ('working_time_id', '!=',  False)])
        start, end = attendance_obj.get_start_end_datetime(today)
        start_time = attendance_obj.str_to_datetime(start) - timedelta(hours=7)
        end_time = attendance_obj.str_to_datetime(end) - timedelta(hours=7)
        for employee_id in employee_ids:
            # Check In
            domain_ci = [('mncei_employee_id', '=', employee_id.id), ('check_in', '>=', start_time), ('check_in', '<=', end_time)]
            attendance_ci = attendance_obj.env['hr.attendance'].sudo().search(domain_ci, order='id desc')
            # Check Out
            domain_co = [('mncei_employee_id', '=', employee_id.id), ('check_out', '>=', start_time), ('check_out', '<=', end_time)]
            attendance_co = attendance_obj.env['hr.attendance'].sudo().search(domain_co, order='id desc')
            # ==================================================================
            merged_set = set(attendance_ci).union(set(attendance_co))
            # Mengonversi semua data
            attendance_ids = list(merged_set)
            if not attendance_ids:
                times = _('%s 00:00:00' % (today))
                attendance_id = attendance_obj.create({
                    'mncei_employee_id': employee_id.id,
                    'type_ci': 'Alpha',
                    'is_alpha': True,
                    'check_in': datetime.strptime(times, "%Y-%m-%d %H:%M:%S") - timedelta(hours=7),
                    'check_out': datetime.strptime(times, "%Y-%m-%d %H:%M:%S") - timedelta(hours=7)
                })
                leave_id = leave_obj.search([('mncei_employee_id', '=', employee_id.id), ('date_from', '>=', start_time), ('date_to', '<=', end_time)])
                if leave_id:
                    attendance_id.write({
                        'is_leave': True,
                        'is_alpha': False,
                    })
                # else:
                #     if employee_id.head_user2:
                #         alpha_date = datetime.strptime(times, "%Y-%m-%d %H:%M:%S")
                #         head_employee = employee_id.head_user2.nama_lengkap
                #         # Send Mail
                #         template_alpha = self.env.ref('mnc_attendance.notification_alpha')
                #         template_alpha.with_context(alpha_date=alpha_date.date(), head_employee=head_employee).send_mail(attendance_id.id, force_send=True, email_values={'email_to': employee_id.head_user2.email})
        return

    @api.onchange('check_in')
    def _change_alpha_flag(self):
        if self.check_in:
            self.is_alpha = False

    def check_presence(self, data_list, key):
        return any(item[0].lower() == key.lower() for item in data_list)

    def query_check_in(self, ttype, department_id):
        query = """
            SELECT emp.nama_lengkap AS emp_name, emp.nip_char AS NIK, dept.name AS Dept, emp.head_user2 AS Head, comp.name AS company, att.check_in + INTERVAL '5 hours' AS CheckIn, att.check_out + INTERVAL '5 hours' AS CheckOut, att.create_date + INTERVAL '7 hours'  AS AttDate, att.type_ci, att.type_co
            FROM hr_attendance att
            JOIN mncei_employee emp ON emp.id = att.mncei_employee_id
            JOIN mncei_department dept ON emp.department = dept.id
            JOIN res_company comp ON emp.company = comp.id
            WHERE att.type_ci in ('%s') AND att.create_date + INTERVAL '7 hours' >= CURRENT_DATE
            AND att.create_date + INTERVAL '7 hours' < CURRENT_DATE + INTERVAL '1 day' AND dept.id = %i
            GROUP BY emp_name, NIK, dept.name, Head, comp.name, CheckIn, CheckOut, AttDate, att.type_ci, att.type_co
        """ % (ttype, department_id)
        return query

    def query_check_out(self, ttype, department_id):
        query = """
            SELECT emp.nama_lengkap AS emp_name, emp.id AS employee_id, emp.nip_char AS NIK, dept.name AS Dept, emp.head_user2 AS Head, comp.name AS company, att.check_in + INTERVAL '5 hours' AS CheckIn, att.check_out + INTERVAL '5 hours' AS CheckOut, att.create_date + INTERVAL '7 hours'  AS AttDate, att.type_ci, att.type_co
            FROM hr_attendance att
            JOIN mncei_employee emp ON emp.id = att.mncei_employee_id
            JOIN mncei_department dept ON emp.department = dept.id
            JOIN res_company comp ON emp.company = comp.id
            WHERE att.type_co in ('%s') AND att.create_date + INTERVAL '7 hours' >= CURRENT_DATE
            AND att.create_date + INTERVAL '7 hours' < CURRENT_DATE + INTERVAL '1 day' AND dept.id = %i
            GROUP BY emp_name, emp.id, NIK, dept.name, Head, comp.name, CheckIn, CheckOut, AttDate, att.type_ci, att.type_co
        """ % (ttype, department_id)
        return query

    def query_attendance(self, ttype, department_id):
        late_in_ci = self.check_presence(LIST_CI, ttype)
        late_in_co = self.check_presence(LIST_CO, ttype)
        # Check 'Late' in LIST_CO
        if late_in_ci:
            query = self.query_check_in(ttype, department_id)
        elif late_in_co:
            query = self.query_check_out(ttype, department_id)
        # Exce
        self.env.cr.execute(query)
        query_result = self.env.cr.dictfetchall()
        # Group By Head
        head_dict = defaultdict(list)
        for emp in query_result:
            if emp['head']:
                head_dict[emp['head']].append(emp)
        # Mengurutkan karyawan berdasarkan checkin dan mengambil tiga teratas untuk setiap head
        top_employees_per_head = {}
        for head, emp_list in head_dict.items():
            if late_in_ci:
                sorted_employees = sorted(emp_list, key=lambda x: x['checkin'], reverse=True)
            elif late_in_co:
                sorted_employees = sorted(emp_list, key=lambda x: x['checkout'], reverse=True)
            top_employees_per_head[head] = sorted_employees[:3]
        return top_employees_per_head

    # Notification Late
    @api.model
    def send_report_attendance(self):
        attendance_obj = self.env['hr.attendance']
        department_ids = self.env['mncei.department'].search([])
        for department_id in department_ids:
            # ============== Late ==============
            late_datas = attendance_obj.query_attendance('Late', department_id)
            # Send Late
            for head, top_employees in late_datas.items():
                employee_head = self.env['mncei.employee'].browse(head)
                if len(top_employees) > 0:
                    template_alpha = self.env.ref('mnc_attendance.notification_late').with_context(dbname=self._cr.dbname, top_employees=top_employees)
                    template_alpha.with_context(top_employees=top_employees).send_mail(employee_head.id, force_send=True, email_values={'email_to': employee_head.email})

    # Notification Resume Member Team
    @api.model
    def send_report_resume_attendance(self):
        attendance_obj = self.env['hr.attendance']
        department_ids = self.env['mncei.department'].search([])
        for department_id in department_ids:
            attendance_obj.query_attendance('Late', department_id)
        return

    # Kirim notifikasi email untuk setiap karyawan yang tidak hadir selama 3 hari berturut-turut
    @api.model
    def check_consecutive_absences_and_notify(self):
        # Ambil semua data kehadiran dengan kondisi 3 Hari kebelakang
        today = fields.Date.today()
        start, end = self.get_start_end_datetime(today)
        start_time = self.str_to_datetime(start) - timedelta(days=1)
        end_time = self.str_to_datetime(start) - timedelta(days=4)
        # Search Record
        attendances = self.search([('type_ci', '=', 'Alpha'), ('check_in', '!=', False), ('check_in', '<=', start_time), ('check_in', '>=', end_time)], order='check_in asc')
        attendance_records = {}
        # Mengelompokkan data kehadiran berdasarkan karyawan
        for record in attendances:
            employee_id = record.mncei_employee_id.id
            check_in_date = fields.Datetime.from_string(record.check_in).date()

            if employee_id not in attendance_records:
                attendance_records[employee_id] = []

            attendance_records[employee_id].append(check_in_date)

        # absentees = self._check_consecutive_absence(attendance_records)
        if attendance_records:
            for employee, dates in attendance_records.items():
                emp = self.env['mncei.employee'].browse(employee)
        return
        # Kirim notifikasi email untuk setiap karyawan yang tidak hadir selama 3 hari berturut-turut
        # for employee_id in absentees:
        #     self._send_absence_notification(employee_id)
